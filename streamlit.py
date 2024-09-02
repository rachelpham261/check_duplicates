import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def process_excel(file_path):
    # Load the original Excel workbook
    original_wb = load_workbook(file_path)
    all_sheets = original_wb.sheetnames

    # Filter and sort sheets based on date format MM.DD.YY
    date_sheets = []
    non_date_sheets = []
    for sheet in all_sheets:
        try:
            parsed_date = datetime.strptime(sheet, '%m.%d.%y')
            date_sheets.append((sheet, parsed_date))
        except ValueError:
            non_date_sheets.append(sheet)

    # Sort sheets by the parsed date
    date_sheets.sort(key=lambda x: x[1])
    sorted_sheets = [sheet[0] for sheet in date_sheets]

    # Initialize an empty dataframe for the summary of unique addresses
    summary_df = pd.DataFrame(columns=[
        'Site Address House Number', 
        'Site Address Street Prefix', 
        'Site Address Street Name', 
        'Site Address Unit Number', 
        'Site Address City', 
        'Site Address State', 
        'Site Address Zip+4',
        'Full Site Address', 
        'First Appearance Date'
    ])

    # Dictionary to keep track of where each address has appeared
    address_appearance = {}

    # Highlight color
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Create a new workbook for the processed file
    processed_file_path = os.path.splitext(file_path)[0] + '_Processed.xlsx'
    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        # Copy all non-date sheets to the new workbook
        for sheet_name in non_date_sheets:
            original_sheet = original_wb[sheet_name]
            writer.book.create_sheet(sheet_name)
            new_sheet = writer.book[sheet_name]
            
            for row in original_sheet.iter_rows(values_only=True):
                new_sheet.append(row)

        # Process and replace date-formatted sheets
        for sheet in sorted_sheets:
            df = pd.read_excel(file_path, sheet_name=sheet)

            # Convert all relevant address columns to string and strip them
            address_columns = [
                'Site Address House Number', 
                'Site Address Street Prefix', 
                'Site Address Street Name', 
                'Site Address Unit Number', 
                'Site Address City', 
                'Site Address State', 
                'Site Address Zip+4'
            ]
            df[address_columns] = df[address_columns].fillna('').astype(str).apply(lambda x: x.str.strip())

            # Concatenate the address components in the desired format
            df['Full Site Address'] = (
                df['Site Address House Number'] + ' ' +
                df['Site Address Street Prefix'] + ' ' + df['Site Address Street Name'] + ', ' +
                df['Site Address Unit Number'].apply(lambda x: f"Unit {x}, " if x != '' else '') +
                df['Site Address City'] + ', ' +
                df['Site Address State'] + ' ' + df['Site Address Zip+4']
            ).str.replace(r'\s+', ' ', regex=True).str.strip().replace(', ,', ',', regex=False)

            # Calculate preceding dates for each full address
            def find_preceding_dates(address):
                preceding_dates = [
                    date for date, addresses in address_appearance.items()
                    if address in addresses
                ]
                return ', '.join(sorted(preceding_dates))

            df['Preceding Dates'] = df['Full Site Address'].apply(find_preceding_dates)

            # Update the address_appearance dictionary
            for address in df['Full Site Address']:
                if sheet not in address_appearance:
                    address_appearance[sheet] = set()
                address_appearance[sheet].add(address)

            # Reorder columns
            columns = address_columns + ['Full Site Address', 'Preceding Dates'] + [col for col in df.columns if col not in address_columns + ['Full Site Address', 'Preceding Dates']]
            df = df[columns]

            # Write the updated sheet, replacing the original
            df.to_excel(writer, sheet_name=sheet, index=False)

            # Apply highlighting
            worksheet = writer.sheets[sheet]
            for col in ['H', 'I']:  # Assuming 'Full Site Address' is column H and 'Preceding Dates' is column I
                for cell in worksheet[col]:
                    cell.fill = highlight_fill

            # Add to summary dataframe
            unique_df = df.drop_duplicates(subset=['Full Site Address'])[['Full Site Address'] + address_columns]
            unique_df['First Appearance Date'] = unique_df['Full Site Address'].apply(lambda x: find_preceding_dates(x).split(', ')[0] if find_preceding_dates(x) else sheet)
            summary_df = pd.concat([summary_df, unique_df], ignore_index=True)

        # Write the summary of unique addresses to a new sheet
        summary_df.drop_duplicates(subset=['Full Site Address']).to_excel(writer, sheet_name='Summary of Unique Addresses', index=False)
        worksheet = writer.sheets['Summary of Unique Addresses']
        for col in ['H', 'I']:  # Assuming 'Full Site Address' is column H and 'First Appearance Date' is column I
            for cell in worksheet[col]:
                cell.fill = highlight_fill

    return processed_file_path

def main():
    st.title("Check for Duplicate Addresses")

    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

    if uploaded_file is not None:
        # Save the uploaded file temporarily
        with open(uploaded_file.name, "wb") as f:
            f.write(uploaded_file.getbuffer())

        st.write("Processing the file...")
        processed_file = process_excel(uploaded_file.name)

        st.write("Download the processed file:")
        with open(processed_file, "rb") as f:
            st.download_button(label="Download", data=f, file_name=os.path.basename(processed_file))

        # Clean up temporary files
        os.remove(uploaded_file.name)
        os.remove(processed_file)

if __name__ == "__main__":
    main()
